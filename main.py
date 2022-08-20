from openpyxl import load_workbook

from json import dump

def isin_list(to_check : str, check_list : list[list[str]]):
    for item in check_list:
        if item[0].strip().lower() in to_check.lower() and item[1].strip().lower() in to_check.lower():
            return True
    return False

wb_name = "timetable.xlsx"

wb = load_workbook(wb_name)

excess_sheets = wb.sheetnames[1:]

for i in excess_sheets:
    wb.remove(wb[i])

ws = wb.worksheets[0]

# col = 1
# for i in range(10):
#     cell = ws.cell(row=2, column=col)
#     if (cell.value == "BDS-2020"):
#         break
#     col += 1

courselist = []
degree = ""

with open("pref.txt", 'r') as f:
    degree = f.readline().strip("\n").strip()
    courselist = [line.split("-") for line in f.readlines()]

timings = ["8:30", "10:00", "11:30", "1:00", "2:30", "4:00", "5:30", "7:00"]

timings_dict = {time:None for time in timings}

days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

ttable = {}
for day in days:
    ttable[day] = timings_dict.copy()

timings_idx = 0
clash_flag = False
for col in range(6, 6+(9*8),9):

    for row in range(5, 245):
        curr_cell = ws.cell(row=row, column=col)
        if (curr_cell.value == None):
            continue
        if (degree.lower() in curr_cell.value.lower() and isin_list(curr_cell.value, courselist)):
            newrow = row
            
            while (str(ws.cell(row=newrow, column=1).value) == "None" and newrow > 0):
                newrow -= 1
            
            day = str(ws.cell(row=newrow, column=1).value)
            day = day.strip().split(" ")[0]
            if (ttable[day][timings[timings_idx]] != None):
                print(f"ERROR: {curr_cell.value} clashing with {ttable[day][timings[timings_idx]]}")
                clash_flag = True
                break

            ttable[day][timings[timings_idx]] = str(curr_cell.value) + f" [VENUE:{str(ws.cell(row=row, column=2).value)}]"
            if ("lab" in str(curr_cell.value).lower()):

                if (ttable[day][timings[timings_idx + 1]] != None):
                    print(f"ERROR: {curr_cell.value} clashing with {ttable[day][timings[timings_idx + 1]]}")
                    clash_flag = True
                    break

                ttable[day][timings[timings_idx + 1]] = str(curr_cell.value) + f" [VENUE:{str(ws.cell(row=row, column=2).value)}]"


        if clash_flag:
            break
        
    timings_idx += 1
        
if not clash_flag:
    dump(ttable, open("TimeTable.json", 'w'))

print("COMPLETED")